"""
Classe base de todos os geradores de relatório.

Cada subclasse define:
- `type_id` / `title` / `template_name` (HTML do Django, herda base.html)
- `fetch_data()` — busca os dados do banco (pode chamar set_progress)
- `build_context(data)` — monta o ctx HTML
- `table_columns()` / `table_rows(data)` — define o formato tabular (XLSX/CSV)
- `filters_display(data)` — lista de filtros aplicados (label + value) p/ header

A base lida com PDF (WeasyPrint), DOCX (htmldocx), XLSX (openpyxl) e CSV
(csv stdlib) usando o template HTML + os métodos table_* das subclasses.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Iterable

from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone

from apps.reports.models import ReportJob


@dataclass(frozen=True)
class TableColumn:
    """Definição de uma coluna em relatório tabular (XLSX/CSV).

    `key` é a chave usada em `table_rows` para extrair o valor.
    `format` é opcional — usado por openpyxl pra formatar células (ex.: 'dd/mm/yyyy').
    `width` é hint visual no XLSX (em "caracteres aproximados" do openpyxl).
    """
    key: str
    label: str
    format: str | None = None
    width: int = 22


@dataclass(frozen=True)
class FilterDisplay:
    """Item da lista de filtros aplicados (rendered no header do PDF/DOCX)."""
    label: str
    value: str


class BaseReport:
    """Classe abstrata. Subclasses sobrescrevem os métodos marcados."""

    type_id: str = ''
    title: str = ''
    subtitle: str | None = None
    template_name: str = ''
    # 'portrait' (default) ou 'landscape' — afeta CSS do PDF.
    orientation: str = 'portrait'

    def __init__(self, job: ReportJob):
        self.job = job
        self.filters: dict[str, Any] = dict(job.filters or {})
        self.user = job.user
        self.include_header = job.include_header

    # ─────────────────── métodos que subclasses sobrescrevem ───────────────────

    def fetch_data(self) -> Any:
        """Busca os dados do banco. Pode chamar self.set_progress(N, msg)."""
        raise NotImplementedError

    def build_context(self, data: Any) -> dict[str, Any]:
        """Contexto extra do template (campos específicos do relatório)."""
        return {}

    def table_columns(self) -> list[TableColumn]:
        """Colunas do XLSX/CSV. Vazio = sem suporte tabular."""
        return []

    def table_rows(self, data: Any) -> Iterable[dict[str, Any]]:
        """Itera linhas no formato dict (chaves = TableColumn.key)."""
        return []

    def filters_display(self, data: Any) -> list[FilterDisplay]:
        """Filtros aplicados, formatados pra exibição."""
        return []

    # ─────────────────── helpers ───────────────────

    def set_progress(self, value: int, message: str = '') -> None:
        """Atualiza progresso no DB (consumido pelo polling do frontend)."""
        value = max(0, min(100, int(value)))
        self.job.progress = value
        if message:
            self.job.progress_message = message[:200]
        self.job.save(update_fields=['progress', 'progress_message', 'updated_at'])

    def paginate_with_progress(
        self,
        queryset,
        *,
        label: str,
        progress_start: int = 10,
        progress_end: int = 65,
        chunk_size: int = 100,
    ) -> list:
        """Materializa um queryset em chunks reportando progresso a cada página.

        - `queryset`: já com filtros, select_related, order_by aplicados.
        - `label`: prefixo da mensagem (ex.: "Exportando cards").
        - `progress_start`/`progress_end`: faixa de progresso (0–100) coberta
          por esta operação. O caller usa o resto da faixa para tarefas
          subsequentes (render PDF, etc.).
        - `chunk_size`: tamanho de cada página interna.

        Reporta no DB a cada chunk: "Exportando cards 245 / 1340". O frontend
        faz polling a cada 1.5s e vê a barra evoluir.

        Retorna a lista completa (consume todos os chunks).
        """
        # COUNT é uma única query — necessária pra calcular % real.
        total = queryset.count()
        if total == 0:
            self.set_progress(progress_end, f'{label} 0 / 0')
            return []

        rows: list = []
        # Loop manual: queryset[start:end] gera SQL com LIMIT/OFFSET. Cada
        # chunk vira UMA query. Pra QS pequenos isso é overhead irrelevante;
        # pra QS grandes (centenas+) o ganho de UX (barra evoluindo) vale.
        progress_span = max(1, progress_end - progress_start)
        for offset in range(0, total, chunk_size):
            chunk = list(queryset[offset:offset + chunk_size])
            rows.extend(chunk)
            loaded = min(offset + chunk_size, total)
            pct = progress_start + int(progress_span * loaded / total)
            self.set_progress(pct, f'{label} {loaded} / {total}')
        return rows

    def filename(self, fmt: str) -> str:
        """Formato padrão: 'BWATech - <Título> - DD-MM-YYYY HH-MM.<fmt>'

        Caracteres inválidos pro filesystem (Windows: \\/:*?\"<>|) são
        substituídos por hífen — title fica legível sem quebrar download.
        """
        import re
        ts = timezone.localtime().strftime('%d-%m-%Y %H-%M')
        raw_title = self.title or self.type_id or 'Relatorio'
        # Substitui caracteres proibidos em nomes de arquivo (Windows é mais
        # restritivo que Linux/macOS — atende todos).
        safe_title = re.sub(r'[\\/:*?"<>|]', '-', raw_title).strip()
        # Colapsa espaços/hífens duplicados pra ficar legível.
        safe_title = re.sub(r'\s+', ' ', safe_title)
        return f'BWATech - {safe_title} - {ts}.{fmt}'

    def _logo_path(self) -> str | None:
        """Caminho absoluto do logo para o WeasyPrint embedar.

        Procura em STATICFILES_DIRS / static / frontend/dist/assets.
        Retorna None se não achar — header usa o fallback de texto.
        """
        import os
        candidates = [
            os.path.join(settings.BASE_DIR, '..', 'frontend', 'public', 'assets', 'bwa-tech-black.png'),
            os.path.join(settings.BASE_DIR, '..', 'frontend', 'dist', 'assets', 'bwa-tech-black.png'),
            os.path.join(settings.BASE_DIR, 'static', 'bwa-tech-black.png'),
        ]
        for path in candidates:
            abs_path = os.path.abspath(path)
            if os.path.exists(abs_path):
                # WeasyPrint precisa de URI file:// para imagens locais.
                return 'file://' + abs_path.replace('\\', '/')
        return None

    def _common_context(self, data: Any) -> dict[str, Any]:
        """Contexto comum (header, meta, filtros) — usado em PDF e DOCX."""
        full_name = (
            f'{(self.user.first_name or "").strip()} {(self.user.last_name or "").strip()}'.strip()
            or self.user.username
        )
        return {
            'report_title': self.title,
            'report_subtitle': self.subtitle,
            'generated_at': timezone.localtime().strftime('%d/%m/%Y %H:%M'),
            'user_name': full_name,
            'filters_display': self.filters_display(data),
            'logo_path': self._logo_path(),
            'orientation': self.orientation,
        }

    def render_html(self, data: Any) -> str:
        ctx = {**self._common_context(data), **self.build_context(data)}
        return render_to_string(self.template_name, ctx)

    # ─────────────────── geração por formato ───────────────────

    def render(self, fmt: str) -> bytes:
        if fmt == 'pdf':
            return self._render_pdf()
        if fmt == 'docx':
            return self._render_docx()
        if fmt == 'xlsx':
            return self._render_xlsx()
        if fmt == 'csv':
            return self._render_csv()
        raise ValueError(f'Formato não suportado: {fmt}')

    def _render_pdf(self) -> bytes:
        """HTML → PDF via WeasyPrint.

        WeasyPrint depende de libs nativas do sistema (Pango, Cairo, GDK-Pixbuf,
        GObject). No Linux Docker (produção) elas vêm via apt-get. No Windows
        é preciso instalar GTK runtime — sem isso o import falha com erro FFI
        genérico que confunde o usuário. Catch específico aqui devolve uma
        mensagem prática.
        """
        try:
            from weasyprint import HTML
        except OSError as exc:
            msg = str(exc)
            if 'libgobject' in msg or 'libpango' in msg or 'libcairo' in msg:
                raise RuntimeError(
                    'Geração de PDF indisponível neste ambiente: faltam as libs '
                    'nativas do WeasyPrint (Pango/Cairo/GObject). '
                    'Em Windows isso é normal — use DOCX, XLSX ou CSV para testes '
                    'locais. Em produção Linux as libs estão instaladas no Dockerfile.'
                ) from exc
            raise

        data = self.fetch_data()
        self.set_progress(80, 'Renderizando PDF...')
        html_str = self.render_html(data)
        # base_url = caminho do projeto pra resolver assets relativos (logo).
        return HTML(string=html_str, base_url=str(settings.BASE_DIR)).write_pdf()

    def _render_docx(self) -> bytes:
        """HTML → DOCX via htmldocx (que escreve no python-docx Document).

        DOCX herda muito menos CSS que PDF — paginação, footers e cores
        de fundo são simplificadas. Bom o suficiente pra um relatório
        editável.
        """
        from docx import Document
        from htmldocx import HtmlToDocx

        data = self.fetch_data()
        self.set_progress(80, 'Renderizando DOCX...')
        html_str = self.render_html(data)

        doc = Document()
        HtmlToDocx().add_html_to_document(html_str, doc)
        buf = io.BytesIO()
        doc.save(buf)
        return buf.getvalue()

    def _render_xlsx(self) -> bytes:
        """Tabela via openpyxl. Primeira linhas são CONTEXTO (sempre); depois
        opcionalmente uma linha de cabeçalho das colunas (se include_header),
        e em seguida os dados.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        data = self.fetch_data()
        cols = self.table_columns()
        if not cols:
            raise NotImplementedError(
                f'Relatório {self.type_id} não suporta exportação tabular (XLSX).'
            )

        self.set_progress(70, 'Montando planilha...')

        wb = Workbook()
        ws = wb.active
        ws.title = (self.title or 'Relatório')[:30]

        # --- Cabeçalho contextual (sempre) ---
        bold = Font(bold=True, color='FFFFFF')
        primary_fill = PatternFill('solid', fgColor='754C99')
        meta_font = Font(bold=True)

        ws.cell(row=1, column=1, value=self.title).font = Font(bold=True, size=14, color='754C99')
        full_name = (
            f'{(self.user.first_name or "").strip()} {(self.user.last_name or "").strip()}'.strip()
            or self.user.username
        )
        ws.cell(row=2, column=1, value='Gerado em:').font = meta_font
        ws.cell(row=2, column=2, value=timezone.localtime().strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=3, column=1, value='Gerado por:').font = meta_font
        ws.cell(row=3, column=2, value=full_name)

        filt_list = self.filters_display(data)
        cursor = 4
        if filt_list:
            ws.cell(row=cursor, column=1, value='Filtros aplicados').font = Font(bold=True, color='754C99')
            cursor += 1
            for f in filt_list:
                ws.cell(row=cursor, column=1, value=f.label).font = meta_font
                ws.cell(row=cursor, column=2, value=f.value)
                cursor += 1
        cursor += 1  # linha em branco

        # --- Linha de cabeçalho das colunas (opcional) ---
        data_start = cursor
        if self.include_header:
            for idx, col in enumerate(cols, start=1):
                cell = ws.cell(row=cursor, column=idx, value=col.label)
                cell.font = bold
                cell.fill = primary_fill
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cursor += 1
            data_start = cursor

        # --- Linhas de dados ---
        total_rows = len(list(self.table_rows(data))) if hasattr(self.table_rows(data), '__len__') else 0
        # Pra evitar materializar duas vezes a sequência: re-iterar
        rows_iter = self.table_rows(data)
        row_count = 0
        for row in rows_iter:
            for idx, col in enumerate(cols, start=1):
                value = row.get(col.key, '')
                cell = ws.cell(row=cursor, column=idx, value=value)
                if col.format:
                    cell.number_format = col.format
            cursor += 1
            row_count += 1
            # Atualiza progresso a cada 50 linhas — evita queries a cada linha.
            if total_rows and row_count % 50 == 0:
                pct = 70 + int(25 * row_count / max(total_rows, 1))
                self.set_progress(pct, f'Exportando linhas {row_count} / {total_rows}')

        # --- Larguras das colunas ---
        for idx, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = col.width

        # Congela cabeçalho das colunas (se houver).
        if self.include_header:
            ws.freeze_panes = ws.cell(row=data_start, column=1)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _render_csv(self) -> bytes:
        """CSV simples. Cabeçalho contextual nas primeiras linhas (comentário),
        depois linha de header (opcional), depois dados.

        Usa ; como delimitador (mais amigável ao Excel em pt-BR) e \\r\\n
        como line terminator (padrão CSV).
        """
        data = self.fetch_data()
        cols = self.table_columns()
        if not cols:
            raise NotImplementedError(
                f'Relatório {self.type_id} não suporta exportação tabular (CSV).'
            )

        self.set_progress(70, 'Montando CSV...')

        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=';', lineterminator='\r\n')

        full_name = (
            f'{(self.user.first_name or "").strip()} {(self.user.last_name or "").strip()}'.strip()
            or self.user.username
        )
        # Linhas de contexto (sempre — independente do include_header).
        writer.writerow([f'# {self.title}'])
        writer.writerow([f'# Gerado em: {timezone.localtime().strftime("%d/%m/%Y %H:%M")}'])
        writer.writerow([f'# Gerado por: {full_name}'])
        for f in self.filters_display(data):
            writer.writerow([f'# {f.label}: {f.value}'])
        writer.writerow([])  # linha em branco

        if self.include_header:
            writer.writerow([col.label for col in cols])

        for row in self.table_rows(data):
            writer.writerow([_csv_cell(row.get(col.key, '')) for col in cols])

        return buf.getvalue().encode('utf-8-sig')  # BOM ajuda Excel pt-BR


def _csv_cell(value: Any) -> str:
    """Normaliza valor para string CSV."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y %H:%M')
    return str(value)
